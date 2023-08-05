import openpyxl
from openpyxl.styles import Font

def write_dataframe_to_excel(dataframe, sheet_name, file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        workbook.create_sheet(sheet_name)

    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)

    sheet = workbook[sheet_name]

    data = [list(dataframe.columns)] + dataframe.values.tolist()

    for row in data:
        sheet.append(row)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    workbook.save(file_path)
